import logging
import os
from telegram import (
    Update,
    InlineKeyboardButton,
    InlineKeyboardMarkup,
)
from telegram.ext import (
    ApplicationBuilder,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
)
from openpyxl import Workbook, load_workbook

# ---------------- CONFIG ----------------
TOKEN = "8459453645:AAFEK1AgNcZOmWD5Hmv1_watrib53PbTkBo"
CHANNEL_LINK = "https://t.me/enocToken"
TWITTER_LINK = "https://x.com/enocToken"
RETWEET_LINK = "https://x.com/enocToken/status/1234567890"
WALLET_FOLDER = os.path.join(os.path.expanduser("~/Desktop"), "wallet")

# Ensure wallet folder exists
os.makedirs(WALLET_FOLDER, exist_ok=True)

# Excel file for storing wallet data
wallet_file = os.path.join(WALLET_FOLDER, "wallet_data.xlsx")

if not os.path.exists(wallet_file):
    wb = Workbook()
    ws = wb.active
    ws.append(["User ID", "Username", "Wallet Address"])
    wb.save(wallet_file)

# ---------------- LOGGING ----------------
logging.basicConfig(format="%(asctime)s - %(name)s - %(levelname)s - %(message)s", level=logging.INFO)

# ---------------- HANDLERS ----------------
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    user_id = user.id

    welcome_text = (
        f"👋 Hello {user.first_name or 'EnocToken'}!\n\n"
        "🎉 Welcome to the ENOC Airdrop campaign!\n\n"
        "💎 How to join:\n"
        f"1️⃣ Join our channel: {CHANNEL_LINK}\n"
        f"2️⃣ Follow our X (Twitter): {TWITTER_LINK}\n"
        f"3️⃣ Retweet our pinned post: {RETWEET_LINK}\n"
        "4️⃣ Register your wallet using: /wallet <youraddress>\n\n"
        "🔥 First 1000 participants join FREE. After that, entry costs $0.1 (simulated).\n\n"
        "🔗 Share your referral link to earn 10 ENOC per successful referral:\n"
        f"https://t.me/ENOCAirdropBot?start=ref{user_id}\n\n"
        "👇 Tap a button below to proceed:"
    )

    keyboard = [
        [
            InlineKeyboardButton("📢 Join ENOC Channel", url=CHANNEL_LINK),
        ],
        [
            InlineKeyboardButton("🐦 Follow ENOC on X", url=TWITTER_LINK),
            InlineKeyboardButton("🔁 Retweet Post", url=RETWEET_LINK),
        ],
        [
            InlineKeyboardButton("💼 Register Wallet", callback_data="register_wallet"),
            InlineKeyboardButton("🔗 Show My Referral Link", callback_data="referral_link"),
        ],
    ]

    reply_markup = InlineKeyboardMarkup(keyboard)
    await update.message.reply_text(welcome_text, reply_markup=reply_markup)


async def wallet(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user = update.effective_user
    if len(context.args) == 0:
        await update.message.reply_text("⚠️ Please use the command like this:\n/wallet YOUR_WALLET_ADDRESS")
        return

    wallet_address = context.args[0]

    # Save to Excel
    wb = load_workbook(wallet_file)
    ws = wb.active
    ws.append([user.id, user.username or "NoUsername", wallet_address])
    wb.save(wallet_file)

    await update.message.reply_text(f"✅ Wallet registered successfully!\nAddress: `{wallet_address}`", parse_mode="Markdown")


async def button_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()

    user = update.effective_user
    user_id = user.id

    if query.data == "register_wallet":
        await query.message.reply_text("💼 Please register your wallet using:\n/wallet <your_wallet_address>")

    elif query.data == "referral_link":
        await query.message.reply_text(
            f"🔗 Your referral link:\nhttps://t.me/ENOCAirdropBot?start=ref{user_id}"
        )


# ---------------- MAIN ----------------
if __name__ == "__main__":
    print("✅ ENOC Airdrop Bot is running...")
    app = ApplicationBuilder().token(TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("wallet", wallet))
    app.add_handler(CallbackQueryHandler(button_callback))

    app.run_polling()